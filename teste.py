import tkinter as tk
from tkinter import filedialog
import openpyxl

def selecionar_arquivo_e_tabela():
    """Abre uma janela para o usuário selecionar um arquivo Excel e uma tabela."""
    global arquivo_selecionado, tabela_selecionada

    # Seleciona o arquivo
    arquivo_selecionado = filedialog.askopenfilename(
        initialdir="/", title="Selecione o arquivo Excel",
        filetypes=(("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*"))
    )

    if arquivo_selecionado:
        # Carrega o arquivo e obtém a lista de tabelas
        workbook = openpyxl.load_workbook(arquivo_selecionado)
        sheetnames = workbook.sheetnames

        # Cria uma nova janela para selecionar a tabela
        janela_tabelas = tk.Toplevel(janela)
        janela_tabelas.title("Selecione a Tabela")

        # Cria uma listbox para exibir as tabelas
        listbox_tabelas = tk.Listbox(janela_tabelas)
        for sheetname in sheetnames:
            listbox_tabelas.insert(tk.END, sheetname)
        listbox_tabelas.pack()

        # Função para obter a tabela selecionada
        def obter_tabela_selecionada():
            tabela_selecionada = listbox_tabelas.get(tk.ACTIVE)
            janela_tabelas.destroy()
            # Chama a função para editar a tabela
            editar_tabela(arquivo_selecionado, tabela_selecionada)

        # Botão para confirmar a seleção da tabela
        botao_confirmar = tk.Button(janela_tabelas, text="Confirmar", command=obter_tabela_selecionada)
        botao_confirmar.pack()

def editar_tabela(arquivo, tabela):
    """Edita a tabela especificada.

    Args:
        arquivo: Caminho completo para o arquivo Excel.
        tabela: Nome da tabela a ser editada.
    """
    workbook = openpyxl.load_workbook(arquivo)
    sheet = workbook[tabela]
    # ... suas operações com a planilha ...
    workbook.save(arquivo)

# ... (resto do código)