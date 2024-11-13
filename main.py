import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import load_workbook
from ttkbootstrap import Style
from ttkbootstrap.constants import *
from datetime import datetime


#variáveis
camposLimpar = []
dadosInseridos = [
    
]
i = 1


# Funções


def limparCampos():
    for campo in camposLimpar:
        campo.delete(0, END)


#def calcularHoraSobreaviso():
#    horaEnt= horaEntrada.get()
#    horaSai = horaSaida.get()
    

# função principal para enviar dados para tabela
def enviarParaTabela():
    diaTrabalhado_get = diaTrabalhado.get()
    horaEntrada_get = int(horaEntrada.get())
    horaSaida_get = int(horaSaida.get())
    sobreaviso = horaSaida_get - horaEntrada_get
    servico_get = servico.get()

    camposLimpar.extend([diaTrabalhado, horaEntrada, horaSaida, servico])
    
    if diaTrabalhado_get and horaEntrada_get and horaSaida_get and servico_get:
        listaDiasInseridos.insert(tk.END, diaTrabalhado_get)
        dadosInseridos.insert(i, [diaTrabalhado_get, f"{horaEntrada_get}:00", f"{horaSaida_get}:00", f"{sobreaviso}:00", servico_get])
        print(dadosInseridos)
        limparCampos()   

    else: 
        messagebox.showwarning("Aviso", "Há campos a serem preenchidos")



def finalizar():
    
    if listaDiasInseridos.size() > 0:    
        row = 7
        coluna = 2
        nome_get = nome.get()
        chapa_get = chapa.get()
        for dado in dadosInseridos:
            for item in dado: 
                planilha.cell(row=row, column=coluna, value=item).number_format = 'General'
                coluna += 1
                print(f'loop item, coluna {coluna}')
            row += 1
            coluna = 2
            print(F"passou loop dado row {row}")
        messagebox.showinfo("Sucesso", "O Din Din cairá logo!")
        planilha['D5'] = f"Nome: {nome_get}"
        planilha['C5'] = chapa_get
        planilhaModelo.save('./programaFacil/teste.xlsx')
        root.destroy()
    else:
        messagebox.showwarning("Aviso", "Não há registros")


# carregar e configurar planilha.

planilhaModelo = load_workbook("./PlanilhaModelo")
planilha = planilhaModelo.active
planilha['D4'] = datetime.today().strftime('%Y')


# geração de configuração da página
root = tk.Tk()
root.title("Din Din (Sobreaviso)")
icon = tk.PhotoImage(file="./programaFacil/icone.png")
root.iconphoto(True, icon) 
style = Style(theme='cyborg')
FonteTitulo = ("Arial", 20, "bold")


# Selecionar table a ser editada



# Credenciais únicas
textoNome = ttk.Label(root, text="Nome: ")
textoNome.grid(row=2, column=1)
nome = ttk.Entry(root)
nome.grid(row=3, column=1)
#############################
textoChapa = ttk.Label(root, text="Chapa: ")
textoChapa.grid(row=2, column=2)
chapa = ttk.Entry(root)
chapa.grid(row=3, column=2)


# Credencias de dias trabalhados
textoDiaTrablhado = ttk.Label(root, text="Dia")
textoDiaTrablhado.grid(row=4, column=0)
diaTrabalhado = ttk.Entry(root)
diaTrabalhado.grid(row=5, column=0)
#
textoHoraEntrada = ttk.Label(root, text="Entrada")
textoHoraEntrada.grid(row=4, column=1)
horaEntrada = ttk.Entry(root)
horaEntrada.grid(row=5, column=1)
#
textHoraSaida = ttk.Label(root, text="Saida")
textHoraSaida.grid(row=4, column=2)
horaSaida = ttk.Entry(root)
horaSaida.grid(row=5, column=2)
#
textoServico = ttk.Label(root, text="Serviço")
textoServico.grid(row=4, column=3)
servico = ttk.Entry(root)
servico.grid(row=5, column=3)



#checkbox para encerrar programa
#Se estiver marcada, encerra o programa 
finalizarPlanilha = ttk.Button(root, text='Finaliza Planilha', command=finalizar)
finalizarPlanilha.grid(row=10, column=0, columnspan=7)

# Botão para enviar as horas
botao = ttk.Button(root, text="Adicionar", command=enviarParaTabela)
botao.grid(row=6, column=0, columnspan=7)


# Listbox para exibir os dados inseridos
listaDiasInseridos = tk.Listbox(root)
listaDiasInseridos.grid(row=8, column=0, columnspan=7, sticky="nsew")


#Avisos


root.mainloop()